"""
PT Rahaza — Sprint 3.1: HR Reports Module

Endpoints (prefix /api/rahaza/hr/reports):
  - GET /attendance-summary?from=&to=&department_id=&location_id=&shift_id=&employee_id=
  - GET /overtime-summary?from=&to=&department_id=&location_id=&shift_id=&employee_id=
  - GET /payroll-summary?period_code=&department_id=&location_id=&shift_id=
  - GET /turnover?from=&to=&department_id=
  - GET /attendance-summary.xlsx (export Excel)
  - GET /attendance-summary.pdf (export PDF)
  - GET /overtime-summary.xlsx
  - GET /payroll-summary.xlsx
  - GET /turnover.xlsx

Sprint 3.1 Goal:
  - Kombinasi table + charts untuk visual insights
  - Export Excel + PDF untuk reporting
  - Filter by department/location/shift untuk granular analysis
"""
from fastapi import APIRouter, Request, HTTPException, Response
from fastapi.responses import StreamingResponse
from database import get_db
from auth import require_auth, serialize_doc
import logging
from datetime import datetime, timezone, date, timedelta
from typing import Optional
import io

log = logging.getLogger(__name__)
router = APIRouter(prefix="/api/rahaza/hr/reports", tags=["rahaza-hr-reports"])


def _now(): return datetime.now(timezone.utc)


async def _require_hr_admin(request: Request):
    """Require HR, manager, or admin access."""
    user = await require_auth(request)
    role = (user.get("role") or "").lower()
    if role in ("superadmin", "admin", "owner", "manager", "hr", "staff_hr", "manager_produksi"):
        return user
    perms = user.get("_permissions") or []
    if "*" in perms or "hr.manage" in perms or "hr.view_reports" in perms:
        return user
    raise HTTPException(403, "Forbidden: butuh akses HR/Manager untuk melihat laporan.")


# ── Attendance Summary Report ──────────────────────────────────────────────────

@router.get("/attendance-summary")
async def get_attendance_summary(
    request: Request,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    department_id: Optional[str] = None,
    location_id: Optional[str] = None,
    shift_id: Optional[str] = None,
    employee_id: Optional[str] = None,
):
    """
    Attendance summary report: per employee/period/dept/location/shift.
    Returns:
      - List of employees with attendance stats
      - Aggregate totals
      - Chart data (daily trends)
    """
    await _require_hr_admin(request)
    db = get_db()
    
    # Default to current month if no dates provided
    if not from_date or not to_date:
        today = date.today()
        from_date = today.replace(day=1).isoformat()
        last_day = (today.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        to_date = last_day.isoformat()
    
    # Build employee filter
    emp_query = {"active": True}
    if department_id:
        emp_query["department_id"] = department_id
    if location_id:
        emp_query["location_id"] = location_id
    if shift_id:
        emp_query["shift_id"] = shift_id
    if employee_id:
        emp_query["id"] = employee_id
    
    employees = await db.rahaza_employees.find(emp_query, {"_id": 0}).to_list(None)
    emp_ids = [e["id"] for e in employees]
    
    if not emp_ids:
        return {
            "summary": [],
            "aggregates": {},
            "chart_data": [],
            "period": {"from": from_date, "to": to_date},
        }
    
    # Fetch attendance events in period
    att_query = {
        "employee_id": {"$in": emp_ids},
        "date": {"$gte": from_date, "$lte": to_date},
    }
    attendance_events = await db.rahaza_attendance_events.find(att_query, {"_id": 0}).to_list(None)
    
    # Calculate working days in period
    d_from = date.fromisoformat(from_date)
    d_to = date.fromisoformat(to_date)
    total_days = (d_to - d_from).days + 1
    
    # Group by employee
    emp_map = {e["id"]: e for e in employees}
    att_by_emp = {}
    for att in attendance_events:
        eid = att["employee_id"]
        if eid not in att_by_emp:
            att_by_emp[eid] = []
        att_by_emp[eid].append(att)
    
    # Build summary per employee
    summary = []
    for emp in employees:
        eid = emp["id"]
        emp_atts = att_by_emp.get(eid, [])
        
        stats = {
            "hadir": 0,
            "izin": 0,
            "sakit": 0,
            "cuti": 0,
            "alfa": 0,      # FIX: use "alfa" not "alpha" (matches attendance status)
            "terlambat": 0,
        }
        
        for att in emp_atts:
            status = (att.get("status") or "").lower()
            if status in ("hadir", "present"):
                stats["hadir"] += 1
                if att.get("is_late"):
                    stats["terlambat"] += 1
            elif status == "izin":
                stats["izin"] += 1
            elif status == "sakit":
                stats["sakit"] += 1
            elif status == "cuti":
                stats["cuti"] += 1
            elif status in ("alfa", "alpha", "absent", "absen"):  # FIX: handle all variants
                stats["alfa"] += 1
        
        recorded_days = len(emp_atts)
        missing_days = total_days - recorded_days
        
        summary.append({
            "employee_id": eid,
            "employee_code": emp.get("employee_code"),
            "employee_name": emp.get("name"),
            "department_name": emp.get("department_name"),
            "location_name": emp.get("location_name"),
            "shift_name": emp.get("shift_name"),
            "total_days": total_days,
            "recorded_days": recorded_days,
            "missing_days": missing_days,
            **stats,
            "attendance_rate": round((stats["hadir"] / total_days * 100), 1) if total_days > 0 else 0,
        })
    
    # Aggregates
    total_employees = len(summary)
    total_hadir = sum(s["hadir"] for s in summary)
    total_izin = sum(s["izin"] for s in summary)
    total_sakit = sum(s["sakit"] for s in summary)
    total_cuti = sum(s["cuti"] for s in summary)
    total_alfa = sum(s["alfa"] for s in summary)
    total_terlambat = sum(s["terlambat"] for s in summary)
    avg_attendance_rate = round(sum(s["attendance_rate"] for s in summary) / total_employees, 1) if total_employees > 0 else 0

    # Chart data: daily attendance trend
    daily_counts = {}
    current = d_from
    while current <= d_to:
        daily_counts[current.isoformat()] = {"hadir": 0, "izin": 0, "sakit": 0, "cuti": 0, "alfa": 0}
        current += timedelta(days=1)

    for att in attendance_events:
        d_key = att.get("date")
        status = (att.get("status") or "").lower()
        if d_key in daily_counts:
            if status in ("hadir", "present"):
                daily_counts[d_key]["hadir"] += 1
            elif status == "izin":
                daily_counts[d_key]["izin"] += 1
            elif status == "sakit":
                daily_counts[d_key]["sakit"] += 1
            elif status == "cuti":
                daily_counts[d_key]["cuti"] += 1
            elif status in ("alfa", "alpha", "absent", "absen"):
                daily_counts[d_key]["alfa"] += 1

    chart_data = [
        {"date": d_key, **counts}
        for d_key, counts in sorted(daily_counts.items())
    ]

    return serialize_doc({
        "summary": summary,
        "aggregates": {
            "total_employees": total_employees,
            "total_records": total_employees * total_days,
            "total_hadir": total_hadir,
            "total_izin": total_izin,
            "total_sakit": total_sakit,
            "total_cuti": total_cuti,
            "total_alfa": total_alfa,
            "alfa_count": total_alfa,
            "total_terlambat": total_terlambat,
            "avg_attendance_rate": avg_attendance_rate,
            "present_pct": avg_attendance_rate,
        },
        "chart_data": chart_data,
        "period": {"from": from_date, "to": to_date},
    })


# ── Overtime Summary Report ────────────────────────────────────────────────────

@router.get("/overtime-summary")
async def get_overtime_summary(
    request: Request,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    department_id: Optional[str] = None,
    location_id: Optional[str] = None,
    shift_id: Optional[str] = None,
    employee_id: Optional[str] = None,
):
    """
    Overtime summary: hours per employee in period.
    Returns:
      - Per employee overtime hours
      - Aggregates (total hours, avg per employee)
      - Chart data (daily/weekly trend)
    """
    await _require_hr_admin(request)
    db = get_db()
    
    if not from_date or not to_date:
        today = date.today()
        from_date = today.replace(day=1).isoformat()
        last_day = (today.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        to_date = last_day.isoformat()
    
    # Build employee filter
    emp_query = {"active": True}
    if department_id:
        emp_query["department_id"] = department_id
    if location_id:
        emp_query["location_id"] = location_id
    if shift_id:
        emp_query["shift_id"] = shift_id
    if employee_id:
        emp_query["id"] = employee_id
    
    employees = await db.rahaza_employees.find(emp_query, {"_id": 0}).to_list(None)
    emp_ids = [e["id"] for e in employees]
    
    if not emp_ids:
        return {
            "summary": [],
            "aggregates": {},
            "chart_data": [],
            "period": {"from": from_date, "to": to_date},
        }
    
    # Fetch attendance with overtime hours
    att_query = {
        "employee_id": {"$in": emp_ids},
        "date": {"$gte": from_date, "$lte": to_date},
        "overtime_hours": {"$exists": True, "$gt": 0},
    }
    overtime_events = await db.rahaza_attendance_events.find(att_query, {"_id": 0}).to_list(None)
    
    # Group by employee
    emp_map = {e["id"]: e for e in employees}
    ot_by_emp = {}
    for ot in overtime_events:
        eid = ot["employee_id"]
        if eid not in ot_by_emp:
            ot_by_emp[eid] = []
        ot_by_emp[eid].append(ot)
    
    # Build summary
    summary = []
    for emp in employees:
        eid = emp["id"]
        ot_records = ot_by_emp.get(eid, [])
        total_ot_hours = sum(float(r.get("overtime_hours") or 0) for r in ot_records)
        ot_days = len(ot_records)
        
        if total_ot_hours > 0 or ot_days > 0:
            summary.append({
                "employee_id": eid,
                "employee_code": emp.get("employee_code"),
                "employee_name": emp.get("name"),
                "department_name": emp.get("department_name"),
                "location_name": emp.get("location_name"),
                "total_ot_hours": round(total_ot_hours, 2),
                "ot_days": ot_days,
                "avg_ot_per_day": round(total_ot_hours / ot_days, 2) if ot_days > 0 else 0,
            })
    
    # Aggregates
    total_employees_with_ot = len(summary)
    total_ot_hours = sum(s["total_ot_hours"] for s in summary)
    avg_ot_per_emp = round(total_ot_hours / total_employees_with_ot, 2) if total_employees_with_ot > 0 else 0
    
    # Chart data: daily OT hours
    d_from = date.fromisoformat(from_date)
    d_to = date.fromisoformat(to_date)
    daily_ot = {}
    current = d_from
    while current <= d_to:
        daily_ot[current.isoformat()] = 0
        current += timedelta(days=1)
    
    for ot in overtime_events:
        d = ot.get("date")
        if d in daily_ot:
            daily_ot[d] += float(ot.get("overtime_hours") or 0)
    
    chart_data = [
        {"date": d, "ot_hours": round(hours, 2)}
        for d, hours in sorted(daily_ot.items())
    ]
    
    return serialize_doc({
        "summary": summary,
        "aggregates": {
            "total_employees_with_ot": total_employees_with_ot,
            "total_ot_hours": total_ot_hours,
            "avg_ot_per_emp": avg_ot_per_emp,
        },
        "chart_data": chart_data,
        "period": {"from": from_date, "to": to_date},
    })


# ── Payroll Summary Report ─────────────────────────────────────────────────────

@router.get("/payroll-summary")
async def get_payroll_summary(
    request: Request,
    period_code: Optional[str] = None,
    department_id: Optional[str] = None,
    location_id: Optional[str] = None,
    shift_id: Optional[str] = None,
):
    """
    Payroll summary: per payroll run / period.
    Returns:
      - Per employee payslip summary
      - Aggregates (total gross, deductions, net)
      - Chart data (breakdown by component)
    """
    await _require_hr_admin(request)
    db = get_db()
    
    # If no period_code, use latest payroll run
    if not period_code:
        latest_run = await db.rahaza_payroll_runs.find_one(
            {"status": {"$in": ["completed", "approved", "finalized"]}},  # FIX: include finalized
            {"_id": 0},
            sort=[("created_at", -1)]
        )
        if not latest_run:
            return {
                "summary": [],
                "aggregates": {},
                "chart_data": [],
                "period": None,
            }
        run_id = latest_run["id"]
    else:
        run = await db.rahaza_payroll_runs.find_one({"period_code": period_code}, {"_id": 0})
        if not run:
            raise HTTPException(404, "Payroll run tidak ditemukan.")
        run_id = run["id"]
    
    # Fetch payslips for this run
    payslip_query = {"run_id": run_id}
    payslips = await db.rahaza_payslips.find(payslip_query, {"_id": 0}).to_list(None)
    
    # Fetch employees for filters
    if department_id or location_id or shift_id:
        emp_ids_filtered = []
        emp_query = {"active": True}
        if department_id:
            emp_query["department_id"] = department_id
        if location_id:
            emp_query["location_id"] = location_id
        if shift_id:
            emp_query["shift_id"] = shift_id
        employees = await db.rahaza_employees.find(emp_query, {"_id": 0, "id": 1}).to_list(None)
        emp_ids_filtered = {e["id"] for e in employees}
        payslips = [p for p in payslips if p.get("employee_id") in emp_ids_filtered]
    
    # Enrich with employee names
    emp_ids = list({p["employee_id"] for p in payslips if p.get("employee_id")})
    employees = await db.rahaza_employees.find({"id": {"$in": emp_ids}}, {"_id": 0}).to_list(None) if emp_ids else []
    emp_map = {e["id"]: e for e in employees}
    
    summary = []
    for p in payslips:
        emp = emp_map.get(p.get("employee_id")) or {}
        # FIX: Support multiple field name conventions (gross_pay from payroll, gross from admin seed)
        gross = p.get("gross_pay") or p.get("gross_salary") or p.get("gross") or 0
        ded   = p.get("deductions_total") or p.get("total_deductions") or 0
        if not ded and isinstance(p.get("deductions"), list):
            ded = sum(d.get("amount", 0) for d in p["deductions"])
        elif not ded and isinstance(p.get("deductions"), dict):
            ded = sum(p["deductions"].values())
        net = p.get("net_pay") or p.get("net_salary") or p.get("net") or max(0, gross - ded)
        summary.append({
            "employee_id": p.get("employee_id"),
            "employee_code": emp.get("employee_code"),
            "employee_name": emp.get("name"),
            "department_name": emp.get("department_name"),
            "gross_salary": gross,
            "total_deductions": ded,
            "net_salary": net,
            "attendance_days": p.get("attendance_days", 0),
            "overtime_hours": p.get("overtime_hours", 0),
        })
    
    # Aggregates
    total_employees = len(summary)
    total_gross = sum(s["gross_salary"] for s in summary)
    total_deductions = sum(s["total_deductions"] for s in summary)
    total_net = sum(s["net_salary"] for s in summary)
    avg_net = round(total_net / total_employees, 2) if total_employees > 0 else 0
    
    # Chart data: breakdown by salary component (simplified)
    chart_data = {
        "gross_salary": total_gross,
        "deductions": total_deductions,
        "net_salary": total_net,
    }
    
    return serialize_doc({
        "summary": summary,
        "aggregates": {
            "total_employees": total_employees,
            "total_gross": total_gross,
            "total_deductions": total_deductions,
            "total_net": total_net,
            "avg_net": avg_net,
        },
        "chart_data": chart_data,
        "period": {"period_code": period_code or "latest", "run_id": run_id},
    })


# ── Turnover Report ────────────────────────────────────────────────────────────

@router.get("/turnover")
async def get_turnover_report(
    request: Request,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    department_id: Optional[str] = None,
):
    """
    Employee turnover: hires & resignations in period.
    Returns:
      - New hires list
      - Resignations list
      - Turnover rate calculation
      - Chart data (monthly trend)
    """
    await _require_hr_admin(request)
    db = get_db()
    
    if not from_date or not to_date:
        today = date.today()
        from_date = today.replace(month=1, day=1).isoformat()
        to_date = today.isoformat()
    
    # Build employee filter
    emp_query = {}
    if department_id:
        emp_query["department_id"] = department_id
    
    # New hires (join_date in period)
    hire_query = {
        **emp_query,
        "join_date": {"$gte": from_date, "$lte": to_date},
    }
    new_hires = await db.rahaza_employees.find(hire_query, {"_id": 0}).to_list(None)
    
    # Resignations (resign_date in period)
    resign_query = {
        **emp_query,
        "resign_date": {"$gte": from_date, "$lte": to_date},
        "active": False,
    }
    resignations = await db.rahaza_employees.find(resign_query, {"_id": 0}).to_list(None)
    
    # Active employees at period start (for turnover rate calculation)
    active_start_query = {
        **emp_query,
        "$or": [
            {"join_date": {"$lt": from_date}},
            {"join_date": None},
        ],
        "$or": [
            {"resign_date": None},
            {"resign_date": {"$gte": from_date}},
        ],
    }
    # Simplified: count all active employees (approx)
    active_employees = await db.rahaza_employees.count_documents({**emp_query, "active": True})
    
    # Turnover rate = (resignations / avg_employees) * 100
    avg_employees = active_employees if active_employees > 0 else 1
    turnover_rate = round((len(resignations) / avg_employees) * 100, 2)
    
    # Chart data: monthly hires & resignations
    # (Simplified: just count by month)
    d_from = date.fromisoformat(from_date)
    d_to = date.fromisoformat(to_date)
    
    monthly_data = {}
    current = d_from.replace(day=1)
    while current <= d_to:
        month_key = current.strftime("%Y-%m")
        monthly_data[month_key] = {"hires": 0, "resignations": 0}
        current = (current.replace(day=28) + timedelta(days=4)).replace(day=1)
    
    for emp in new_hires:
        if emp.get("join_date"):
            month_key = emp["join_date"][:7]  # YYYY-MM
            if month_key in monthly_data:
                monthly_data[month_key]["hires"] += 1
    
    for emp in resignations:
        if emp.get("resign_date"):
            month_key = emp["resign_date"][:7]
            if month_key in monthly_data:
                monthly_data[month_key]["resignations"] += 1
    
    chart_data = [
        {"month": m, **counts}
        for m, counts in sorted(monthly_data.items())
    ]
    
    return serialize_doc({
        "new_hires": new_hires,
        "resignations": resignations,
        "aggregates": {
            "new_hires_count": len(new_hires),
            "resignations_count": len(resignations),
            "active_employees": active_employees,
            "turnover_rate": turnover_rate,
        },
        "chart_data": chart_data,
        "period": {"from": from_date, "to": to_date},
    })


# ── Export Endpoints (Excel) ───────────────────────────────────────────────────

@router.get("/attendance-summary.xlsx")
async def export_attendance_excel(
    request: Request,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    department_id: Optional[str] = None,
    location_id: Optional[str] = None,
    shift_id: Optional[str] = None,
    employee_id: Optional[str] = None,
):
    """Export attendance summary to Excel.
    Phase 24E: format tanggal DD/MM/YYYY di header period."""
    # Re-use attendance summary logic
    data = await get_attendance_summary(request, from_date, to_date, department_id, location_id, shift_id, employee_id)
    
    # Phase 24E: helper format DD/MM/YYYY
    def _fmt_id(iso_str):
        if not iso_str: return "-"
        s = str(iso_str)[:10]
        try:
            from datetime import datetime as _dt
            d = _dt.strptime(s, "%Y-%m-%d")
            return d.strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            return s

    # Generate Excel using openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Summary"
        
        # Header (Phase 24E: DD/MM/YYYY format)
        ws.append(["PT Rahaza — Laporan Ringkasan Absensi"])
        ws.append([f"Periode: {_fmt_id(data['period']['from'])} s/d {_fmt_id(data['period']['to'])}"])
        ws.append([])
        
        # Column headers
        headers = ["Kode Karyawan", "Nama Karyawan", "Departemen", "Total Hari", "Hadir", "Izin", "Sakit", "Cuti", "Alpha", "Terlambat", "Tingkat Hadir (%)"]
        ws.append(headers)
        
        # Data rows
        for row in data["summary"]:
            ws.append([
                row.get("employee_code"),
                row.get("employee_name"),
                row.get("department_name"),
                row.get("total_days"),
                row.get("hadir"),
                row.get("izin"),
                row.get("sakit"),
                row.get("cuti"),
                row.get("alpha"),
                row.get("terlambat"),
                row.get("attendance_rate"),
            ])
        
        # Aggregates
        ws.append([])
        agg = data["aggregates"]
        ws.append(["Total Karyawan:", agg["total_employees"]])
        ws.append(["Rata-rata Tingkat Hadir:", agg["avg_attendance_rate"]])
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=attendance_summary_{data['period']['from']}_{data['period']['to']}.xlsx"}
        )
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Cannot generate Excel.")


@router.get("/overtime-summary.xlsx")
async def export_overtime_excel(
    request: Request,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    department_id: Optional[str] = None,
    location_id: Optional[str] = None,
    shift_id: Optional[str] = None,
    employee_id: Optional[str] = None,
):
    """Export overtime summary to Excel."""
    data = await get_overtime_summary(request, from_date, to_date, department_id, location_id, shift_id, employee_id)
    
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Overtime Summary"
        
        ws.append(["PT Rahaza — Overtime Summary Report"])
        ws.append([f"Period: {data['period']['from']} to {data['period']['to']}"])
        ws.append([])
        
        headers = ["Employee Code", "Employee Name", "Department", "Total OT Hours", "OT Days", "Avg OT/Day"]
        ws.append(headers)
        
        for row in data["summary"]:
            ws.append([
                row.get("employee_code"),
                row.get("employee_name"),
                row.get("department_name"),
                row.get("total_ot_hours"),
                row.get("ot_days"),
                row.get("avg_ot_per_day"),
            ])
        
        ws.append([])
        agg = data["aggregates"]
        ws.append(["Total Employees with OT:", agg["total_employees_with_ot"]])
        ws.append(["Total OT Hours:", agg["total_ot_hours"]])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=overtime_summary_{data['period']['from']}_{data['period']['to']}.xlsx"}
        )
    except ImportError:
        raise HTTPException(500, "openpyxl not installed.")


@router.get("/payroll-summary.xlsx")
async def export_payroll_excel(
    request: Request,
    period_code: Optional[str] = None,
    department_id: Optional[str] = None,
    location_id: Optional[str] = None,
    shift_id: Optional[str] = None,
):
    """Export payroll summary to Excel."""
    data = await get_payroll_summary(request, period_code, department_id, location_id, shift_id)
    
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll Summary"
        
        ws.append(["PT Rahaza — Payroll Summary Report"])
        ws.append([f"Period: {data['period'].get('period_code', 'N/A')}"])
        ws.append([])
        
        headers = ["Employee Code", "Employee Name", "Department", "Gross Salary", "Deductions", "Net Salary", "Attendance Days", "OT Hours"]
        ws.append(headers)
        
        for row in data["summary"]:
            ws.append([
                row.get("employee_code"),
                row.get("employee_name"),
                row.get("department_name"),
                row.get("gross_salary"),
                row.get("total_deductions"),
                row.get("net_salary"),
                row.get("attendance_days"),
                row.get("overtime_hours"),
            ])
        
        ws.append([])
        agg = data["aggregates"]
        ws.append(["Total Employees:", agg["total_employees"]])
        ws.append(["Total Gross:", agg["total_gross"]])
        ws.append(["Total Net:", agg["total_net"]])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=payroll_summary_{data['period'].get('period_code', 'latest')}.xlsx"}
        )
    except ImportError:
        raise HTTPException(500, "openpyxl not installed.")


@router.get("/turnover.xlsx")
async def export_turnover_excel(
    request: Request,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    department_id: Optional[str] = None,
):
    """Export turnover report to Excel."""
    data = await get_turnover_report(request, from_date, to_date, department_id)
    
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        
        # Sheet 1: New Hires
        ws1 = wb.active
        ws1.title = "New Hires"
        ws1.append(["Employee Code", "Name", "Department", "Join Date"])
        for emp in data["new_hires"]:
            ws1.append([emp.get("employee_code"), emp.get("name"), emp.get("department_name"), emp.get("join_date")])
        
        # Sheet 2: Resignations
        ws2 = wb.create_sheet("Resignations")
        ws2.append(["Employee Code", "Name", "Department", "Resign Date"])
        for emp in data["resignations"]:
            ws2.append([emp.get("employee_code"), emp.get("name"), emp.get("department_name"), emp.get("resign_date")])
        
        # Sheet 3: Summary
        ws3 = wb.create_sheet("Summary")
        agg = data["aggregates"]
        ws3.append(["Metric", "Value"])
        ws3.append(["New Hires", agg["new_hires_count"]])
        ws3.append(["Resignations", agg["resignations_count"]])
        ws3.append(["Active Employees", agg["active_employees"]])
        ws3.append(["Turnover Rate (%)", agg["turnover_rate"]])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=turnover_report_{data['period']['from']}_{data['period']['to']}.xlsx"}
        )
    except ImportError:
        raise HTTPException(500, "openpyxl not installed.")



# ── Sprint 3.3: Attendance Validation for Payroll ──────────────────────────────

@router.get("/attendance-validation")
async def validate_attendance_for_payroll(
    request: Request,
    period_from: str,
    period_to: str,
    employee_ids: Optional[str] = None,
):
    """
    Validate attendance completeness for payroll period.
    Returns warnings for employees with missing/incomplete attendance.
    
    Query params:
      - period_from: YYYY-MM-DD
      - period_to: YYYY-MM-DD
      - employee_ids: comma-separated employee IDs (optional)
    
    Returns:
      - warnings: list of employees with attendance issues
      - summary: total employees, total warnings
      - attendance_stats: per employee attendance count
    """
    await _require_hr_admin(request)
    db = get_db()
    
    # Parse period
    try:
        d_from = date.fromisoformat(period_from)
        d_to = date.fromisoformat(period_to)
        total_days = (d_to - d_from).days + 1
    except ValueError:
        raise HTTPException(400, "Invalid date format. Use YYYY-MM-DD.")
    
    # Get employees
    emp_query = {"active": True}
    if employee_ids:
        emp_ids_list = [eid.strip() for eid in employee_ids.split(',')]
        emp_query["id"] = {"$in": emp_ids_list}
    
    employees = await db.rahaza_employees.find(emp_query, {"_id": 0}).to_list(None)
    
    if not employees:
        return {
            "warnings": [],
            "summary": {"total_employees": 0, "total_warnings": 0, "warning_rate": 0},
            "attendance_stats": [],
        }
    
    emp_ids = [e["id"] for e in employees]
    
    # Fetch attendance events in period
    att_query = {
        "employee_id": {"$in": emp_ids},
        "date": {"$gte": period_from, "$lte": period_to},
    }
    attendance_events = await db.rahaza_attendance_events.find(att_query, {"_id": 0}).to_list(None)
    
    # Group by employee
    att_by_emp = {}
    for att in attendance_events:
        eid = att["employee_id"]
        if eid not in att_by_emp:
            att_by_emp[eid] = []
        att_by_emp[eid].append(att)
    
    # Build warnings and stats
    warnings = []
    attendance_stats = []
    
    for emp in employees:
        eid = emp["id"]
        emp_atts = att_by_emp.get(eid, [])
        recorded_days = len(emp_atts)
        missing_days = total_days - recorded_days
        
        # Calculate attendance breakdown
        hadir_count = sum(1 for a in emp_atts if (a.get("status") or "").lower() in ("hadir", "present"))
        
        # Build stat
        stat = {
            "employee_id": eid,
            "employee_code": emp.get("employee_code"),
            "employee_name": emp.get("name"),
            "total_days": total_days,
            "recorded_days": recorded_days,
            "missing_days": missing_days,
            "hadir_count": hadir_count,
            "attendance_rate": round((hadir_count / total_days * 100), 1) if total_days > 0 else 0,
        }
        attendance_stats.append(stat)
        
        # Generate warnings
        if missing_days > 0:
            warnings.append({
                **stat,
                "severity": "high" if missing_days >= (total_days * 0.3) else "medium",
                "warning_message": f"{missing_days} hari belum ada record attendance ({round((missing_days/total_days)*100, 1)}% dari periode)",
            })
        elif hadir_count < (total_days * 0.5):
            # Warning jika kehadiran <50% meskipun ada record
            warnings.append({
                **stat,
                "severity": "medium",
                "warning_message": f"Tingkat kehadiran rendah: {stat['attendance_rate']}% (hadir {hadir_count}/{total_days} hari)",
            })
    
    # Sort warnings by severity
    warnings.sort(key=lambda w: (0 if w["severity"] == "high" else 1, w["employee_name"]))
    
    return serialize_doc({
        "warnings": warnings,
        "summary": {
            "total_employees": len(employees),
            "total_warnings": len(warnings),
            "warning_rate": round((len(warnings) / len(employees) * 100), 1) if employees else 0,
            "period": {"from": period_from, "to": period_to, "total_days": total_days},
        },
        "attendance_stats": attendance_stats,
    })
